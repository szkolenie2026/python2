import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font

def formatuj_raport_pro(sciezka_pliku, wybrany_kolor):
    nazwa_arkusza = f"Raport_{wybrany_kolor.capitalize()}"
    
    # KROK 1: Zapisujemy dane i ZAMYKAMY writer
    with pd.ExcelWriter(sciezka_pliku, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_kolor.to_excel(writer, sheet_name=nazwa_arkusza, index=False)
        df_stat.to_excel(writer, sheet_name=nazwa_arkusza, index=True, startcol=13)
    
    # KROK 2: Otwieramy plik ponownie tylko do formatowania
    wb = load_workbook(sciezka_pliku)
    if nazwa_arkusza not in wb.sheetnames:
        print(f"Błąd: Arkusz {nazwa_arkusza} nadal nie istnieje!")
        return
    
    ws = wb[nazwa_arkusza]

    # --- TUTAJ TWOJE FORMATOWANIE ---
    # Tabela
    ostatni_wiersz = len(df_kolor) + 1
    zakres = f"A1:L{ostatni_wiersz}" # Zakładając, że masz 12 kolumn (A-L)
    tab = Table(displayName=f"Tabela_{wybrany_kolor.capitalize()}", ref=zakres)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    
    # Usuwamy stare tabele o tej samej nazwie, jeśli istnieją (zapobiega błędom)
    if tab.displayName in ws.tables:
        del ws.tables[tab.displayName]
        
    ws.add_table(tab)

    # Formatowanie liczb (Kolumna P)
    for row in range(2, 5):
        cell = ws.cell(row=row, column=16)
        cell.number_format = '# ##0.00'
        cell.alignment = Alignment(horizontal='right')

    # Autofit
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # KROK 3: Zapisujemy finalnie
    wb.save(sciezka_pliku)
    print(f"Raport {nazwa_arkusza} sformatowany pomyślnie!")


wybrany_kolor = "BiAłY"
wybrany_kolor = wybrany_kolor.lower()
kurs_euro = 4.30  
df = pd.read_excel('dane.xlsx', sheet_name=0)

# 2. Filtrujemy dane tylko dla wybranego koloru
df_kolor = df[df['Kolor'].str.lower() == wybrany_kolor].copy()

if df_kolor.empty:
    print(f"Brak pojazdów o kolorze: {wybrany_kolor}")
    exit()
# 3. Przeliczamy cenę na Euro i dodajemy nową kolumnę
df_kolor['Cena EUR'] = (df_kolor['Cena netto'] / kurs_euro).round(2)

# 5. Obliczenia
suma_pln = df_kolor['Cena netto'].sum()
srednia_pln = df_kolor['Cena netto'].mean()
liczba_pojazdow = len(df_kolor)
# 6. Tworzymy słownik
dane_stat = {
        "Opis": ["Liczba pojazdów", "Suma (PLN)", "Średnia (PLN)"],
        "Wartość": [
            liczba_pojazdow, 
            suma_pln, 
            round(srednia_pln, 2)
        ]
    }
df_stat = pd.DataFrame(dane_stat)
df_stat.index = range(1, len(df_stat) + 1)
df_stat.index.name = "Lp"


# Wywołanie
formatuj_raport_pro('dane.xlsx', wybrany_kolor)